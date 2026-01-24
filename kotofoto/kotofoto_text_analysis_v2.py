#!/usr/bin/env python3
"""
================================================================================
              АНАЛИЗ КАННИБАЛИЗАЦИИ SEO ТЕКСТОВ ИЗ EXCEL
                  Поиск дублей и похожих описаний товаров
================================================================================

🎯 НАЗНАЧЕНИЕ СКРИПТА:
   Читает kotofoto_text_analysis.xlsx (выход из kotofoto_text_analysis.py)
   → Вычисляет похожесть текстов (TF-IDF + Jaccard)
   → Выявляет дубли и каннибализацию с цветным кодированием
   → Экспортирует результаты в XLSX/CSV/TXT для ревью

📋 ВХОДНЫЕ ДАННЫЕ:
   kotofoto_text_analysis.xlsx (в папке скрипта):
   └─ Столбцы: URL | Текст

📤 ВЫХОДНЫЕ ДАННЫЕ (все с базовым именем kotofoto_text_analysis_v2_result):
   ├─ kotofoto_text_analysis_v2_result.xlsx (пары с цветным кодированием)
   ├─ kotofoto_text_analysis_v2_result.csv (для быстрого анализа)
   └─ kotofoto_text_analysis_v2_result.txt (статистика)

================================================================================
"""

import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==================== КОНФИГ ПУТЕЙ ====================
SCRIPT_DIR = Path(__file__).parent
INPUT_EXCEL = SCRIPT_DIR / "kotofoto_text_analysis.xlsx"
RESULT_BASE = SCRIPT_DIR / "kotofoto_text_analysis_v2_result"
OUTPUT_XLSX = RESULT_BASE.with_suffix(".xlsx")
OUTPUT_CSV = RESULT_BASE.with_suffix(".csv")
OUTPUT_STATS = RESULT_BASE.with_suffix(".txt")

# ==================== КОНФИГ АНАЛИЗА ====================
MIN_SIMILARITY = 0.60
RISK_THRESHOLD_YELLOW = 0.75
RISK_THRESHOLD_RED = 0.80

# ==================== ФУНКЦИИ ====================

def preprocess_text(text):
    """Очистка текста для Jaccard"""
    if not isinstance(text, str):
        return []
    text = text.lower()
    text = "".join(c for c in text if c.isalnum() or c.isspace())
    return text.split()

def jaccard_similarity(text1, text2):
    """Jaccard: пересечение уникальных слов / объединение"""
    words1 = set(preprocess_text(text1))
    words2 = set(preprocess_text(text2))
    if not words1 and not words2:
        return 0.0
    intersection = len(words1.intersection(words2))
    union = len(words1.union(words2))
    return intersection / union if union > 0 else 0.0

def get_risk_color(similarity):
    """Определяет уровень риска и цвет"""
    if similarity >= RISK_THRESHOLD_RED:
        return "🔴 КАННИБАЛИЗАЦИЯ", "red"
    elif similarity >= RISK_THRESHOLD_YELLOW:
        return "🟡 ОЧЕНЬ ПОХОЖИ", "yellow"
    else:
        return "🟢 НОРМА", "green"

def load_and_analyze_excel(filepath):
    """Читает Excel, вычисляет similarity"""
    print(f"📂 Читаем {filepath}...")
    df = pd.read_excel(filepath, sheet_name="Товары")
    
    if "текст" not in df.columns or "url" not in df.columns:
        print("❌ Столбцы 'url' и 'текст' не найдены!")
        print(f"Доступные столбцы: {df.columns.tolist()}")
        exit(1)
    
    texts = df["текст"].fillna("").tolist()
    urls = df["url"].fillna("").tolist()
    
    print(f"   Товаров: {len(texts)}")
    
    if len(texts) < 2:
        print("❌ Нужно минимум 2 товара для сравнения!")
        exit(1)
    
    print(f"\n🔍 Вычисляю TF-IDF similarity...")
    
    vectorizer = TfidfVectorizer(max_features=1000, stop_words=None, lowercase=True)
    try:
        tfidf_matrix = vectorizer.fit_transform(texts)
        tfidf_sim = cosine_similarity(tfidf_matrix)
    except Exception as e:
        print(f"⚠️  TF-IDF ошибка: {e}")
        tfidf_sim = np.zeros((len(texts), len(texts)))
    
    pairs = []
    for i in range(len(texts)):
        for j in range(i + 1, len(texts)):
            tfidf_score = tfidf_sim[i, j]
            jaccard_score = jaccard_similarity(texts[i], texts[j])
            max_score = max(tfidf_score, jaccard_score)
            
            if max_score >= MIN_SIMILARITY:
                risk_text, risk_color = get_risk_color(max_score)
                pairs.append({
                    "url_1": urls[i],
                    "url_2": urls[j],
                    "tfidf_similarity": f"{tfidf_score:.1%}",
                    "jaccard_similarity": f"{jaccard_score:.1%}",
                    "max_similarity": max_score,
                    "риск": risk_text,
                    "color": risk_color
                })
    
    return pd.DataFrame(pairs), tfidf_sim, texts, urls

def color_risk_cells(excel_file, color_col="риск"):
    """Раскрашивает ячейки в Excel по логике цвета"""
    wb = load_workbook(excel_file)
    ws = wb.active
    
    header_row = 1
    risk_col_idx = None
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value == color_col:
            risk_col_idx = col_idx
            break
    
    if risk_col_idx is None:
        print(f"⚠️  Столбец '{color_col}' не найден в Excel")
        return
    
    color_map = {
        "🟢 НОРМА": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "🟡 ОЧЕНЬ ПОХОЖИ": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "🔴 КАННИБАЛИЗАЦИЯ": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        risk_cell = row[risk_col_idx - 1]
        risk_text = risk_cell.value
        
        if risk_text in color_map:
            fill = color_map[risk_text]
            for cell in row:
                cell.fill = fill
    
    wb.save(excel_file)

# ==================== ОСНОВНОЙ КОД ====================
if __name__ == "__main__":
    
    if not INPUT_EXCEL.exists():
        print(f"❌ Файл не найден: {INPUT_EXCEL}")
        print("   Сначала запустите: python kotofoto_text_analysis.py")
        exit(1)
    
    print(f"✅ Нашёл файл: {INPUT_EXCEL}\n")
    
    pairs_df, tfidf_matrix, texts, urls = load_and_analyze_excel(INPUT_EXCEL)
    
    print(f"\n📊 Найдено пар: {len(pairs_df)}")
    
    if len(pairs_df) == 0:
        print("✅ Каннибализация НЕ выявлена (отлично!)")
    else:
        green = len(pairs_df[pairs_df["риск"] == "🟢 НОРМА"])
        yellow = len(pairs_df[pairs_df["риск"] == "🟡 ОЧЕНЬ ПОХОЖИ"])
        red = len(pairs_df[pairs_df["риск"] == "🔴 КАННИБАЛИЗАЦИЯ"])
        
        print(f"🟢 НОРМА (60-75%): {green}")
        print(f"🟡 ОЧЕНЬ ПОХОЖИ (75-80%): {yellow}")
        print(f"🔴 КАННИБАЛИЗАЦИЯ (>80%): {red}")
        
        if red > 0:
            print("\n🔴 ТОП КРИТИЧЕСКИХ ДУБЛЕЙ:")
            risky = pairs_df[pairs_df["риск"] == "🔴 КАННИБАЛИЗАЦИЯ"].sort_values("max_similarity", ascending=False)
            for idx, row in risky.head(3).iterrows():
                print(f"\n   • {row['url_1'][:65]}...")
                print(f"     ↔ {row['url_2'][:65]}...")
                print(f"     TF-IDF: {row['tfidf_similarity']} | Jaccard: {row['jaccard_similarity']}")
        
        pairs_df_sorted = pairs_df.sort_values("max_similarity", ascending=False)
        
        pairs_df_sorted[["url_1", "url_2", "tfidf_similarity", "jaccard_similarity", "риск"]].to_csv(
            OUTPUT_CSV, index=False, encoding="utf-8-sig")
        
        pairs_df_sorted[["url_1", "url_2", "tfidf_similarity", "jaccard_similarity", "риск"]].to_excel(
            OUTPUT_XLSX, sheet_name="Анализ", index=False, engine="openpyxl")
        
        color_risk_cells(OUTPUT_XLSX, color_col="риск")
        
        print(f"\n💾 Экспортировано в папку скрипта:")
        print(f"   ✓ {OUTPUT_XLSX.name}")
        print(f"   ✓ {OUTPUT_CSV.name}")
        
        with open(OUTPUT_STATS, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("               ОТЧЁТ АНАЛИЗА КАННИБАЛИЗАЦИИ\n")
            f.write("=" * 70 + "\n\n")
            f.write(f"Дата анализа: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Входной файл: {INPUT_EXCEL.name}\n\n")
            f.write("СТАТИСТИКА:\n")
            f.write("-" * 70 + "\n")
            f.write(f"Всего товаров: {len(texts)}\n")
            f.write(f"Найдено похожих пар (>60%): {len(pairs_df)}\n\n")
            f.write(f"🟢 НОРМА (60-75%): {green} пар\n")
            f.write(f"🟡 ОЧЕНЬ ПОХОЖИ (75-80%): {yellow} пар\n")
            f.write(f"🔴 КАННИБАЛИЗАЦИЯ (>80%): {red} пар\n\n")
            f.write(f"Средняя похожесть: {pairs_df['max_similarity'].mean():.1%}\n")
            f.write(f"Макс похожесть: {pairs_df['max_similarity'].max():.1%}\n\n")
            f.write("ТОП-5 ДУБЛЕЙ:\n")
            f.write("-" * 70 + "\n")
            for idx, row in pairs_df_sorted.head(5).iterrows():
                f.write(f"\n{idx + 1}. {row['url_1']}\n")
                f.write(f"   ↔ {row['url_2']}\n")
                f.write(f"   TF-IDF: {row['tfidf_similarity']} | Jaccard: {row['jaccard_similarity']}\n")
                f.write(f"   Статус: {row['риск']}\n")
        
        print(f"   ✓ {OUTPUT_STATS.name}\n")

print("✅ Готово!")
