import requests
import time
import pandas as pd
from tqdm import tqdm
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException
from urllib.robotparser import RobotFileParser
from urllib.parse import urlparse
from bs4 import BeautifulSoup
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Reduced list of key Yandex User-Agents (focus on indexing and useful bots)
yandex_agents = [
    "Mozilla/5.0 (compatible; YandexBot/3.0; +http://yandex.com/bots)",
    "Mozilla/5.0 (compatible; YandexImages/3.0; +http://yandex.com/bots)",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 8_1 like Mac OS X) AppleWebKit/600.1.4 (KHTML, like Gecko) Version/8.0 Mobile/12B411 Safari/600.1.4 (compatible; YandexMobileBot/3.0; +http://yandex.com/bots)",
    "Mozilla/5.0 (compatible; YandexMetrika/2.0; +http://yandex.com/bots)",
    "Mozilla/5.0 (compatible; YandexWebmaster/2.0; +http://yandex.com/bots)",
    "Mozilla/5.0 (compatible; YandexFavicons/1.0; +http://yandex.com/bots)",
    "Mozilla/5.0 (compatible; YandexNews/4.0; +http://yandex.com/bots)"
]

# Reduced list of key Google User-Agents (focus on indexing and useful bots)
google_agents = [
    "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",
    "Googlebot-Image/1.0",
    "Googlebot-Video/1.0",
    "Mozilla/5.0 (compatible; Google-InspectionTool/1.0;)"
]

# List of known AI bot User-Agents (compiled from various sources as of 2025)
ai_agents = [
    "Mozilla/5.0 (compatible; AI2Bot/1.0; +https://allenai.org/)",
    "Amazonbot/0.1 (+https://developer.amazon.com/support/amazonbot)",
    "Mozilla/5.0 (compatible; Andibot/1.0)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Claude-Web/1.0; +https://www.anthropic.com)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Anthropic-Claude/1.0; +https://www.anthropic.com)",
    "ClaudeBot/1.0; +https://www.anthropic.com",
    "Mozilla/5.0 (compatible; Applebot-Extended/1.0)",
    "Mozilla/5.0 (compatible; bigsur.ai/1.0)",
    "Mozilla/5.0 (compatible; Brightbot/1.0)",
    "Mozilla/5.0 (compatible; Bytespider; [[email protected]])",
    "TerraCotta https://github.com/CeramicTeam/CeramicTerracotta",
    "Mozilla/5.0 (compatible; Character-AI/1.0; +https://character.ai/)",
    "Mozilla/5.0 (compatible; Devin/1.0)",
    "Mozilla/5.0 (compatible; Cohere-AI/1.0; +https://cohere.com/)",
    "Mozilla/5.0 (compatible; Cohere-Command/1.0; +https://cohere.com/)",
    "CCBot/2.0[](https://commoncrawl.org/faq/)",
    "Mozilla/5.0 (compatible; Crawlspace/1.0)",
    "Mozilla/5.0 (compatible; DeepseekBot/1.0; +https://www.deepseek.com/bot)",
    "Mozilla/5.0 (compatible; Diffbot/0.1; +http://www.diffbot.com/our-apis/crawler/)",
    "Mozilla/5.0 (compatible; DuckAssistBot/1.0; +https://duckduckgo.com/duckassist)",
    "Mozilla/5.0 (compatible; FirecrawlAgent/1.0)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Bard-AI/1.0; +https://developers.google.com/search/docs/crawling-indexing/google-common-crawlers)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Gemini-AI/1.0; +https://developers.google.com/search/docs/crawling-indexing/google-common-crawlers)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Gemini-Deep-Research/1.0)",
    "Mozilla/5.0 (compatible; Google-CloudVertexBot/1.0; +https://cloud.google.com/vertex-ai)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Google-Extended/1.0; +https://developers.google.com/search/docs/crawling-indexing/google-common-crawlers)",
    "Mozilla/5.0 (compatible; Google-NotebookLM/1.0; +https://notebooklm.google.com/)",
    "Mozilla/5.0 (compatible; Groq-Bot/1.0; +https://groq.com/)",
    "Mozilla/5.0 (compatible; HuggingFace-Bot/1.0; +https://huggingface.co/)",
    "Mozilla/5.0 (compatible; IbouBot/1.0; [[email protected]]; +https://ibou.io/iboubot.html)",
    "facebookexternalhit/1.1 (+http://www.facebook.com/externalhit_uatext.php)",
    "Meta-ExternalAgent/1.0 (+https://developers.facebook.com/docs/sharing/bot)",
    "meta-webindexer/1.1",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm) Chrome/W.X.Y.Z Safari/537.36 Edg/W.X.Y.Z",
    "MistralAI-User/1.0",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; ChatGPT-Browser/1.0; +https://openai.com/bot)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; ChatGPT-User/1.0; +https://openai.com/bot)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; ChatGPT-User/2.0; +https://openai.com/bot)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; GPTBot/1.0; +https://openai.com/gptbot)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; OAI-SearchBot/1.0; +https://openai.com/searchbot)",
    "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Perplexity-User/1.0; +https://perplexity.ai/bot)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; PerplexityBot/1.0; +https://perplexity.ai/bot)",
    "Mozilla/5.0 (compatible; Replicate-Bot/1.0; +https://replicate.com/)",
    "Mozilla/5.0 (compatible; RunPod-Bot/1.0; +https://runpod.io/)",
    "Mozilla/5.0 (compatible; ImagesiftBot/1.0)",
    "Mozilla/5.0 (compatible; TimpiBot/1.0)",
    "Mozilla/5.0 (compatible; Together-Bot/1.0; +https://together.ai/)",
    "Mozilla/5.0 (compatible; Kangaroo Bot/1.0)",
    "Mozilla/5.0 (compatible; PanguBot/1.0)",
    "Mozilla/5.0 (compatible; Cotoyogi/1.0)",
    "Mozilla/5.0 (compatible; Webzio-Extended/1.0)",
    "Mozilla/5.0 (compatible; xAI-Bot/1.0; +https://x.ai/)",
    "Mozilla/5.0 (compatible; YouBot/1.0; +https://you.com/bot)",
    "Mozilla/5.0 (compatible; anthropic-ai/1.0; +http://www.anthropic.com/bot.html)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko); compatible; ClaudeBot/1.0; +claudebot@anthropic.com)",
    "Mozilla/5.0 (compatible; claude-web/1.0; +http://www.anthropic.com/bot.html)",
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; PerplexityBot/1.0; +https://perplexity.ai/perplexitybot)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/600.2.5 (KHTML, like Gecko) Version/8.0.2 Safari/600.2.5 (Amazonbot/0.1; +https://developer.amazon.com/support/amazonbot)",
    "Mozilla/5.0 (compatible; Applebot/1.0; +http://www.apple.com/bot.html)",
    "Mozilla/5.0 (compatible; Applebot-Extended/1.0; +http://www.apple.com/bot.html)",
    "Mozilla/5.0 (compatible; FacebookBot/1.0; +http://www.facebook.com/bot.html)",
    "Mozilla/5.0 (compatible; meta-externalagent/1.1 (+https://developers.facebook.com/docs/sharing/webmasters/crawler))",
    "LinkedInBot/1.0 (compatible; Mozilla/5.0; Jakarta Commons-HttpClient/3.1 +http://www.linkedin.com)",
    "Mozilla/5.0 (compatible; Bytespider/1.0; +http://www.bytedance.com/bot.html)",
    "Mozilla/5.0 (compatible; DuckAssistBot/1.0; +http://www.duckduckgo.com/bot.html)",
    "Mozilla/5.0 (compatible; cohere-ai/1.0; +http://www.cohere.ai/bot.html)",
    "Mozilla/5.0 (compatible; AI2Bot/1.0; +http://www.allenai.org/crawler)",
    "Mozilla/5.0 (compatible; CCBot/1.0; +http://www.commoncrawl.org/bot.html)",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.1.2) Gecko/20090729 Firefox/3.5.2 (.NET CLR 3.5.30729; Diffbot/0.1; +http://www.diffbot.com)",
    "Mozilla/5.0 (compatible; omgili/1.0; +http://www.omgili.com/bot.html)",
    "Timpibot/0.8 (+http://www.timpi.io)",
    "Mozilla/5.0 (compatible; YouBot (+http://www.you.com))",
    "Mozilla/5.0 (compatible; MistralAI-User/1.0; +https://mistral.ai/bot)"
]

# Add a regular browser User-Agent
regular_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36"

# Combine all agents (unique)
all_agents = list(set(yandex_agents + google_agents + ai_agents + [regular_agent]))

# Define the 5 main bots for special rendering (PC and Mobile)
main_bots = [
    "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",  # Google
    "Mozilla/5.0 (compatible; YandexBot/3.0; +http://yandex.com/bots)",  # Yandex
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm) Chrome/W.X.Y.Z Safari/537.36 Edg/W.X.Y.Z",  # Bing
    "Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; GPTBot/1.0; +https://openai.com/gptbot)",  # ChatGPT (GPTBot)
    "Mozilla/5.0 (compatible; xAI-Bot/1.0; +https://x.ai/)"  # Grok (xAI)
]

# Bot name mapping for screenshot filenames
bot_names = {
    main_bots[0]: "google",
    main_bots[1]: "yandex",
    main_bots[2]: "bing",
    main_bots[3]: "chatgpt",
    main_bots[4]: "grok"
}

def is_forbidden_by_robots(content):
    forbidden_directives = ['noindex', 'nofollow', 'none', 'noarchive', 'nosnippet', 'notranslate', 'noimageindex']
    for directive in forbidden_directives:
        if directive in content.lower():
            return True
    return False

def check_url_availability(url, user_agent):
    headers = {"User-Agent": user_agent}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        status = response.status_code
        available = status == 200
        has_content = available and len(response.text) > 0

        x_robots_tag = response.headers.get('X-Robots-Tag', 'N/A')
        x_forbidden = is_forbidden_by_robots(x_robots_tag)

        meta_content = 'N/A'
        meta_forbidden = False
        if available:
            soup = BeautifulSoup(response.text, 'html.parser')
            meta_robots = soup.find('meta', attrs={'name': 'robots'})
            if meta_robots and 'content' in meta_robots.attrs:
                meta_content = meta_robots['content']
                meta_forbidden = is_forbidden_by_robots(meta_content)

        return available, status, has_content, x_robots_tag, x_forbidden, meta_content, meta_forbidden
    except requests.RequestException as e:
        return False, str(e), False, 'N/A', False, 'N/A', False

def render_page_desktop(url, user_agent, url_part, bot_name):
    options = Options()
    options.add_argument("--headless")
    options.add_argument(f"user-agent={user_agent}")
    try:
        driver = webdriver.Chrome(options=options)
        driver.get(url)
        screenshot_path = f"screenshot_{url_part}_{bot_name}_desktop.png"
        driver.save_screenshot(screenshot_path)
        driver.quit()
        return True, screenshot_path
    except WebDriverException as e:
        return False, str(e)

def render_page_mobile(url, user_agent, url_part, bot_name):
    options = Options()
    options.add_argument("--headless")
    options.add_argument(f"user-agent={user_agent}")
    options.add_experimental_option("mobileEmulation", {"deviceName": "iPhone X"})
    try:
        driver = webdriver.Chrome(options=options)
        driver.get(url)
        screenshot_path = f"screenshot_{url_part}_{bot_name}_mobile.png"
        driver.save_screenshot(screenshot_path)
        driver.quit()
        return True, screenshot_path
    except WebDriverException as e:
        return False, str(e)

def apply_formatting_to_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    # Определяем индексы столбцов (начиная с 1)
    headers = [cell.value for cell in ws[1]]
    robots_allowed_col = headers.index("Robots.txt Allowed") + 1
    x_forbidden_col = headers.index("X-Robots Forbidden") + 1
    meta_forbidden_col = headers.index("Meta Forbidden") + 1

    # Стили: красный жирный шрифт и заполнение фона (светло-красный для читаемости)
    red_bold_font = Font(color="FF0000", bold=True)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Проходим по строкам (начиная со 2-й, после заголовка)
    for row in range(2, ws.max_row + 1):
        robots_allowed = ws.cell(row, robots_allowed_col).value
        x_forbidden = ws.cell(row, x_forbidden_col).value
        meta_forbidden = ws.cell(row, meta_forbidden_col).value

        # Проверяем наличие запрета
        has_forbidden = (
            (robots_allowed == "No") or
            (x_forbidden == "Yes") or
            (meta_forbidden == "Yes")
        )

        if has_forbidden:
            # Применяем стиль ко всей строке
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.font = red_bold_font
                cell.fill = red_fill

    wb.save(filename)

def main():
    url = input("Enter the URL to check: ").strip()
    
    # Ask for rendering option
    do_render = input("Perform rendering (screenshots)? (Y/N): ").strip().upper() == 'Y'
    
    results = []
    available_count = 0
    unavailable_count = 0
    desktop_rendered_count = 0
    mobile_rendered_count = 0
    robots_allowed_count = 0
    x_forbidden_count = 0
    meta_forbidden_count = 0

    # Fetch and parse robots.txt once
    parsed_url = urlparse(url)
    robots_url = f"{parsed_url.scheme}://{parsed_url.netloc}/robots.txt"
    rp = RobotFileParser()
    try:
        robots_response = requests.get(robots_url, timeout=10)
        if robots_response.status_code == 200:
            rp.parse(robots_response.text.splitlines())
        else:
            rp = None  # No robots.txt or error
    except requests.RequestException:
        rp = None

    # Create URL part for filenames (similar to xlsx but without 'check-')
    domain = parsed_url.netloc.replace('.', '-')
    path = parsed_url.path.replace('/', '-').strip('-')
    url_part = f"{domain}-{path}" if path else domain
    if parsed_url.query:
        query = parsed_url.query.replace('&', '-').replace('=', '-').strip('-')
        url_part = f"{url_part}-{query}"

    print(f"Checking availability and robots directives for {len(all_agents)} User-Agents...")
    if do_render:
        print("Rendering enabled (PC and Mobile for 5 main bots).")
    else:
        print("Rendering disabled.")

    for i, agent in enumerate(tqdm(all_agents)):
        available, status, has_content, x_robots_tag, x_forbidden, meta_content, meta_forbidden = check_url_availability(url, agent)

        robots_allowed = 'N/A'
        if rp:
            robots_allowed = rp.can_fetch(agent, url)
        else:
            robots_allowed = True  # Assume allowed if no robots.txt

        desktop_render_success = False
        desktop_render_info = "N/A"
        mobile_render_success = False
        mobile_render_info = "N/A"
        
        if available and do_render and agent in main_bots:
            bot_name = bot_names.get(agent, "unknown")
            desktop_render_success, desktop_render_info = render_page_desktop(url, agent, url_part, bot_name)
            if desktop_render_success:
                desktop_rendered_count += 1
            mobile_render_success, mobile_render_info = render_page_mobile(url, agent, url_part, bot_name)
            if mobile_render_success:
                mobile_rendered_count += 1

        if available:
            available_count += 1
        else:
            unavailable_count += 1

        if robots_allowed:
            robots_allowed_count += 1

        if x_forbidden:
            x_forbidden_count += 1

        if meta_forbidden:
            meta_forbidden_count += 1
        
        results.append({
            "User-Agent": agent,
            "Available": "Yes" if available else "No",
            "Status": status,
            "Has Content": "Yes" if has_content else "No",
            "Robots.txt Allowed": "Yes" if robots_allowed else "No",
            "X-Robots-Tag": x_robots_tag,
            "X-Robots Forbidden": "Yes" if x_forbidden else "No",
            "Meta Robots": meta_content,
            "Meta Forbidden": "Yes" if meta_forbidden else "No",
            "Desktop Rendered": "Yes" if desktop_render_success else "No",
            "Desktop Screenshot": desktop_render_info if desktop_render_success else desktop_render_info,
            "Mobile Rendered": "Yes" if mobile_render_success else "No",
            "Mobile Screenshot": mobile_render_info if mobile_render_success else mobile_render_info
        })
        time.sleep(1)  # 1 second delay between requests

    print("\nSummary:")
    print(f"Available: {available_count}")
    print(f"Unavailable: {unavailable_count}")
    print(f"Robots.txt Allowed: {robots_allowed_count}")
    print(f"X-Robots Forbidden: {x_forbidden_count}")
    print(f"Meta Forbidden: {meta_forbidden_count}")
    if do_render:
        print(f"Successfully Rendered Desktop: {desktop_rendered_count}")
        print(f"Successfully Rendered Mobile: {mobile_rendered_count}")

    # Create dynamic Excel filename based on URL
    filename = f"check-{url_part}.xlsx"

    df = pd.DataFrame(results)
    df.to_excel(filename, index=False)

    # Apply formatting for forbidden rows
    apply_formatting_to_excel(filename)

    print(f"Report saved to {filename} with formatting for forbidden rows.")
    if do_render:
        print("Screenshots saved as screenshot_{url_part}_{bot_name}_desktop.png and screenshot_{url_part}_{bot_name}_mobile.png for main bots.")

if __name__ == "__main__":
    main()