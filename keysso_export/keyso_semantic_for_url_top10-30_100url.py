#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт выгрузки семантики из Keys.so по URL
(Яндекс Москва + Google Москва в один Excel)

Что делает скрипт (логика):
1. Берёт список URL из CSV-файла input_urls.csv (столбец "url").
2. Для КАЖДОГО URL дважды обращается к API Keys.so:
   - база BASE_YANDEX = "msk"  → Яндекс, регион Москва;
   - база BASE_GOOGLE = "gmsk" → Google, регион Москва. [web:231]
3. По каждому URL запрашивает органические ключевые слова (отчёт
   /report/simple/organic/keywords/bypage), постранично, с ограничениями:
   - максимум MAX_PAGES_PER_URL страниц;
   - максимум PER_PAGE ключей на страницу;
   - не уходим дальше позиции 50 по запросу. [file:229]
4. Собирает по каждому ключу:
   - URL страницы;
   - запрос;
   - позицию в выдаче ("Поз.");
   - Wordstat, "!Wordstat", KEI, количество слов;
   - количество документов в выдаче, дата среза, дата выгрузки.
5. Создаёт один Excel-файл вида keywords_by_urls_DD-MM-YY.xlsx с 4 вкладками:
   - "Яндекс МСК — все"   → все ключи из базы msk;
   - "Яндекс МСК 11-29"   → только запросы с позициями 11–29 (зона для роста);
   - "Google МСК — все"   → все ключи из базы gmsk;
   - "Google МСК 11-29"   → только запросы с позициями 11–29.
   На всех вкладках одинаковая структура колонок. [file:229]

Основной сценарий использования:
- Быстро собрать «боевую» семантику по конкретным посадочным страницам;
- Посмотреть, какие запросы уже в ТОП‑10 / ТОП‑30 по Яндексу и Google Москва;
- Использовать вкладки 11–29 для приоритизации точечных доработок.

---------------------------------------------
КАК ПОЛЬЗОВАТЬСЯ СКРИПТОМ (Пошаговая инструкция)
---------------------------------------------

1. Подготовка окружения
   - Нужен установленный Python 3.8+.
   - Установить зависимости:
     pip install requests openpyxl
   - Убедитесь, что у вас есть рабочий API‑ключ Keys.so с доступом к базам
     msk (Яндекс МСК) и gmsk (Google МСК). [web:231]

2. Настроить API-ключ и пути
   - В начале скрипта в переменной API_KEY пропишите свой токен Keys.so.
   - При необходимости измените путь к CSV-файлу в блоке:
       urls = load_urls_from_csv(r"...\input_urls.csv")
   - CSV должен содержать колонку "url", в каждой строке — один адрес страницы.

3. Подготовить входной CSV
   - Создайте файл input_urls.csv в UTF‑8.
   - Первая строка (заголовок): url
   - Далее по строке на каждый URL, например:
       url
       https://site.ru/catalog/tovary/
       https://site.ru/uslugi/seo/

4. Запуск скрипта
   - Откройте терминал/PowerShell в папке со скриптом.
   - Выполните:
       python keyso_semantic_for_url_top10-30_100url.py
     (или py ..., в зависимости от установки Python).
   - Скрипт:
     * определит количество URL (обрежет до MAX_URLS_PER_RUN, если нужно);
     * для каждого URL выгрузит данные по Яндексу МСК;
     * затем по тем же URL выгрузит данные по Google МСК;
     * по окончании сохранит Excel с 4 вкладками в той же папке. [file:229]

5. Интерпретация результата
   - "Яндекс МСК — все" / "Google МСК — все":
     полный список ключей, которые Keys.so видит за страницей.
   - "Яндекс МСК 11-29" / "Google МСК 11-29":
     запросы с позициями 11–29 — идеальные цели для локальных правок:
       * доработать title/H1;
       * усилить релевантность текста;
       * улучшить внутреннюю перелинковку;
       * проверить коммерческие и поведенческие факторы.
   - Колонка "Дата_выгрузки" позволяет отследить, когда именно делался срез.

6. Ограничения и аккуратное использование
   - Скрипт адаптирован под аккуратное использование тарифа Keys.so:
     * PER_PAGE, MAX_PAGES_PER_URL, MAX_URLS_PER_RUN и REQUEST_DELAY
       можно увеличивать/уменьшать под свой лимит.
   - При HTTP 429 (Too Many Requests) скрипт автоматически ждёт 60 секунд
     и повторяет запрос.
   - Не меняйте BASE_YANDEX / BASE_GOOGLE на несуществующие базы,
     иначе получите ошибки от API. [web:231]

Если нужно добавить другие регионы или базы (например, РФ целиком,
СПб или отдельные Google‑базы), достаточно добавить новые константы
BASE_... и аналогичные блоки выгрузки/вкладки по образцу Яндекс/Google МСК.
"""

import csv
import urllib.parse
import requests
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ========= НАСТРОЙКИ API И БАЗ =========

API_KEY = "69424a9d42ab34.64613057787b24c68dcaacc0112242bc4bfe2090"  # ваш токен Keys.so
BASE_URL = "https://api.keys.so"

# Базы Keys.so (региональные базы см. в документации Keys.so). [web:231]
BASE_YANDEX = "msk"   # Яндекс Москва
BASE_GOOGLE = "gmsk"  # Google Москва (google-база)

# ========= ЛИМИТЫ И ТЕХНИЧЕСКИЕ ПАРАМЕТРЫ =========

PER_PAGE = 200            # строк за запрос (per_page в API)
MAX_PAGES_PER_URL = 2     # максимум страниц API на 1 URL
MAX_URLS_PER_RUN = 50     # до 50 URL за запуск
REQUEST_DELAY = 1.5       # пауза между запросами, сек

REPORT_DATE_STR = datetime.today().strftime("%d-%m-%y")

# Сессия requests с заголовком авторизации
session = requests.Session()
session.headers.update({
    "X-Keyso-TOKEN": API_KEY,
    "Accept": "application/json",
})


def get_keywords_by_url(full_url: str, base: str):
    """
    Вытаскивает ключевые слова для одного URL из указанной базы Keys.so.

    :param full_url: полный URL страницы (https://site.ru/path/)
    :param base: код базы Keys.so (например, 'msk' или 'gmsk'). [web:231]
    :return: список словарей с данными по ключам.
    """
    parsed = urllib.parse.urlparse(full_url)
    domain = parsed.netloc.replace("www.", "")
    page_url = parsed.path or "/"
    if parsed.query:
        page_url += "?" + parsed.query

    all_keywords = []

    for page in range(1, MAX_PAGES_PER_URL + 1):
        params = {
            "base": base,
            "domain": domain,
            "page_url": page_url,
            "sort": "pos|asc",
            "page": page,
            "per_page": PER_PAGE,
        }

        url = f"{BASE_URL}/report/simple/organic/keywords/bypage"

        # Мягкая обработка 429: если тариф не выдерживает, даём передохнуть.
        while True:
            resp = session.get(url, params=params)
            if resp.status_code == 429:
                print("⚠ 429 Too Many Requests. Ждём 60 сек и пробуем ещё раз...")
                time.sleep(60)
                continue
            resp.raise_for_status()
            break

        data = resp.json()
        rows = data.get("data", [])

        # Пауза между запросами даже при пустом ответе,
        # чтобы не долбить API слишком часто.
        time.sleep(REQUEST_DELAY)

        if not rows:
            # Больше страниц нет — выходим из цикла по page.
            break

        for row in rows:
            pos = row.get("pos")
            all_keywords.append({
                "URL": full_url,
                "Запрос": row.get("word"),
                "Поз.": pos,
                "Wordstat": row.get("ws"),
                '"!Wordstat"': row.get("wsk"),
                "KEI": row.get("kei"),
                "Слов": row.get("numwords"),
                "Документов": row.get("docs"),
                "Дата": row.get("serpf"),
                "Дата_выгрузки": REPORT_DATE_STR,
            })

        # Если уже ушли далеко по позициям или вернулось меньше PER_PAGE,
        # то либо хвост, либо лимит — дальше по этому URL не идём.
        max_pos = max(r.get("pos") or 9999 for r in rows)
        if max_pos > 50 or len(rows) < PER_PAGE:
            break

    return all_keywords


def write_sheet(ws, rows):
    """
    Записывает список словарей rows на лист ws и оформляет шапку.
    """
    fieldnames = [
        "URL", "Запрос", "Поз.", "Wordstat", '"!Wordstat"',
        "KEI", "Слов", "Документов", "Дата", "Дата_выгрузки",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Шапка
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field))

    # Базовые ширины колонок под SEO-отчёт
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15


def save_to_excel(all_rows_yandex, all_rows_google):
    """
    Создаёт один Excel с 4 вкладками:
    - Яндекс МСК — все
    - Яндекс МСК 11-29
    - Google МСК — все
    - Google МСК 11-29
    """
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Яндекс МСК — все ключи
    ws_y_all = wb.create_sheet("Яндекс МСК — все", 0)
    write_sheet(ws_y_all, all_rows_yandex)

    # Яндекс МСК — позиции 11–29 (зона роста)
    y_11_29 = [
        r for r in all_rows_yandex
        if isinstance(r.get("Поз."), int) and 11 <= r["Поз."] <= 29
    ]
    ws_y_11_29 = wb.create_sheet("Яндекс МСК 11-29", 1)
    write_sheet(ws_y_11_29, y_11_29)

    # Google МСК — все ключи
    ws_g_all = wb.create_sheet("Google МСК — все", 2)
    write_sheet(ws_g_all, all_rows_google)

    # Google МСК — позиции 11–29
    g_11_29 = [
        r for r in all_rows_google
        if isinstance(r.get("Поз."), int) and 11 <= r["Поз."] <= 29
    ]
    ws_g_11_29 = wb.create_sheet("Google МСК 11-29", 3)
    write_sheet(ws_g_11_29, g_11_29)

    filename = f"keywords_by_urls_{REPORT_DATE_STR}.xlsx"
    wb.save(filename)
    print(f"✅ Файл сохранён: {filename}")


def load_urls_from_csv(path: str):
    """
    Загружает URL из CSV с колонкой 'url'.
    Возвращает список строк. [file:229]
    """
    urls = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = (row.get("url") or "").strip()
            if url:
                urls.append(url)
    return urls


if __name__ == "__main__":
    # Путь к входному CSV. При необходимости изменить под свой.
    csv_path = r"C:\Users\fbesp\SCRYPTS\keysso_export\input_urls.csv"

    urls = load_urls_from_csv(csv_path)
    if not urls:
        print("В input_urls.csv не найдено ни одного URL.")
        raise SystemExit(0)

    # Ограничиваем количество URL за один прогон, чтобы беречь тариф.
    urls = urls[:MAX_URLS_PER_RUN]
    print(f"Будет обработано URL: {len(urls)} (из лимита {MAX_URLS_PER_RUN} за запуск)")

    # -------- Яндекс Москва --------
    print("\n=== Яндекс МСК ===")
    all_rows_yandex = []
    for i, url in enumerate(urls, start=1):
        print(f"[Y {i}/{len(urls)}] Обработка: {url}")
        rows = get_keywords_by_url(url, BASE_YANDEX)
        print(f"  Найдено ключей (Яндекс): {len(rows)}")
        all_rows_yandex.extend(rows)

    # -------- Google Москва --------
    print("\n=== Google МСК ===")
    all_rows_google = []
    for i, url in enumerate(urls, start=1):
        print(f"[G {i}/{len(urls)}] Обработка: {url}")
        rows = get_keywords_by_url(url, BASE_GOOGLE)
        print(f"  Найдено ключей (Google): {len(rows)}")
        all_rows_google.extend(rows)

    if not all_rows_yandex and not all_rows_google:
        print("По всем URL не найдено ни одного ключа.")
        raise SystemExit(0)

    save_to_excel(all_rows_yandex, all_rows_google)
