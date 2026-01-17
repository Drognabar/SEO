"""
Скрипт выгружает ключевые фразы для конкретного URL из API Keys.so
и сохраняет их в Excel‑файл с двумя листами.

Что делает:
- Разбирает переданный URL на домен и путь (учитывает параметры ?utm= и т.п.).
- Запрашивает у Keys.so отчет /report/simple/organic/keywords/bypage
  с заданными базой (по умолчанию msk), лимитом строк и числом страниц.
- Собирает все ключевые фразы с полями:
  «Запрос», «Поз.», «Wordstat», «"!Wordstat"», «KEI», «Слов», «Документов», «Дата».
- Отфильтровывает отдельный список фраз, где позиция в выдаче от 11 до 29.
- Создаёт Excel‑файл keywords_by_url.xlsx:
  Лист «Все ключи» — полный список,
  Лист «Позиции 11-29» — только запросы с позицией 11–29.
- Настраивает шапку таблицы (цвет, жирный шрифт, выравнивание) и ширину колонок.
- В конце печатает в консоль количество всех запросов и запросов 11–29, а также имя файла.

Как подготовить:
1. Установить зависимости:
   pip install requests openpyxl
2. Получить свой API‑ключ Keys.so и подставить его в переменную API_KEY.
3. При необходимости изменить базу (base='msk') и лимиты per_page, max_pages.

Как пользоваться:
1. Внизу файла поменять test_url на нужный адрес страницы:
   test_url = "https://пример.ру/страница/"
2. Запустить скрипт из консоли из папки со скриптом:
   py parser_url_keywords.py
3. Дождаться сообщения «✅ Готово!» и открыть файл keywords_by_url.xlsx
   в Excel для анализа запросов и позиций.
"""

import urllib.parse
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

API_KEY = "69424a9d42ab34.64613057787b24c68dcaacc0112242bc4bfe2090"
BASE_URL = "https://api.keys.so"

session = requests.Session()
session.headers.update({
    "X-Keyso-TOKEN": API_KEY,
    "Accept": "application/json",
})

def get_keywords_by_url(full_url: str, base: str = "msk",
                        per_page: int = 500, max_pages: int = 50):
    parsed = urllib.parse.urlparse(full_url)
    domain = parsed.netloc.replace("www.", "")
    page_url = parsed.path or "/"
    if parsed.query:
        page_url += "?" + parsed.query

    all_keywords = []

    for page in range(1, max_pages + 1):
        params = {
            "base": base,
            "domain": domain,
            "page_url": page_url,
            "sort": "pos|asc",
            "page": page,
            "per_page": per_page,
        }

        resp = session.get(
            f"{BASE_URL}/report/simple/organic/keywords/bypage",
            params=params,
        )
        resp.raise_for_status()
        data = resp.json()
        rows = data.get("data", [])

        if not rows:
            break

        for row in rows:
            all_keywords.append({
                "Запрос": row.get("word"),
                "Поз.": row.get("pos"),
                "Wordstat": row.get("ws"),
                '"!Wordstat"': row.get("wsk"),
                "KEI": row.get("kei"),
                "Слов": row.get("numwords"),
                "Документов": row.get("docs"),
                "Дата": row.get("serpf"),
            })

        if len(rows) < per_page:
            break

    return all_keywords

def save_keywords_to_excel(all_kws, filtered_kws, filename):
    wb = Workbook()
    
    # Удаляем лист по умолчанию
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Лист 1: все ключи
    ws_all = wb.create_sheet("Все ключи", 0)
    write_sheet(ws_all, all_kws)
    
    # Лист 2: ключи с позицией 11-29
    ws_filtered = wb.create_sheet("Позиции 11-29", 1)
    write_sheet(ws_filtered, filtered_kws)
    
    wb.save(filename)

def write_sheet(ws, rows):
    fieldnames = [
        "Запрос", "Поз.", "Wordstat", '"!Wordstat"',
        "KEI", "Слов", "Документов", "Дата",
    ]
    
    # Заголовок
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = row.get(field)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

if __name__ == "__main__":
    test_url = "https://periodica.press/products/photobooks/"

    kws = get_keywords_by_url(test_url)

    # Фильтруем: позиция > 10 и < 30
    kws_11_29 = [
        row for row in kws
        if isinstance(row.get("Поз."), int) and 10 < row["Поз."] < 30
    ]

    # Сохраняем в Excel с двумя листами
    save_keywords_to_excel(kws, kws_11_29, "keywords_by_url.xlsx")

    print("✅ Готово!")
    print(f"  Всего ключей: {len(kws)}")
    print(f"  Ключей с позицией 11–29: {len(kws_11_29)}")
    print(f"  Файл: keywords_by_url.xlsx")
